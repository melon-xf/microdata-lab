from __future__ import annotations

import csv
import math
import re
import warnings
from pathlib import Path
from urllib.parse import unquote, urljoin, urlparse

import duckdb
import httpx
import numpy as np
import pandas as pd
import yaml
from openpyxl import load_workbook
from pydantic import AnyHttpUrl, BaseModel, ConfigDict, TypeAdapter
from selectolax.parser import HTMLParser

from microdata_lab.adapters.base import SourceAdapter
from microdata_lab.bls import build_bls_client, require_bls_response
from microdata_lab.models import (
    DiscoveredArtifact,
    DiscoveredRelease,
    StoredArtifact,
    ValidationResult,
)

INTERVIEW_PRIOR = "interview_prior"
INTERVIEW_CURRENT = "interview_current"
DIARY_CURRENT = "diary_current"
DICTIONARY = "dictionary"
GETTING_STARTED = "getting_started"
INTEGRATION_WORKBOOK = "integration_workbook"
BENCHMARK_WORKBOOK = "benchmark_workbook"
SAMPLE_R = "sample_r"
PRIOR_ERRATA = "prior_errata"
_DEFINITION_PATH = Path(__file__).resolve().parents[3] / "config/ce/2024.yaml"
_url_adapter = TypeAdapter(AnyHttpUrl)


class CEStaticArtifact(BaseModel):
    model_config = ConfigDict(extra="forbid")

    role: str
    url: AnyHttpUrl
    filename: str
    documentation: bool = False


class CEMemberContract(BaseModel):
    model_config = ConfigDict(extra="forbid")

    role: str
    path: str
    family: str
    rows: int
    columns: int
    quarter: str


class CEWeightContract(BaseModel):
    model_config = ConfigDict(extra="forbid")

    full_sample: str
    replicate_prefix: str
    replicate_count: int
    variance_factor: float


class CEBenchmark(BaseModel):
    model_config = ConfigDict(extra="forbid")

    ucc: str
    label: str
    source_code: str
    reference_year: int
    expected_mean: float
    expected_standard_error: float
    expected_population_thousands: float
    expected_positive_reporters: int
    rounding_tolerance: float
    population_rounding_tolerance: float
    workbook_sheet: str
    workbook_label: str


class CEDictionaryContract(BaseModel):
    model_config = ConfigDict(extra="forbid")

    sheet: str
    required_variables: list[str]


class CEReleaseDefinition(BaseModel):
    model_config = ConfigDict(extra="forbid")

    year: int
    landing_page: AnyHttpUrl
    required_artifacts: list[str]
    static_artifacts: list[CEStaticArtifact]
    members: list[CEMemberContract]
    weight_contract: CEWeightContract
    benchmark: CEBenchmark
    dictionary: CEDictionaryContract


class CEAdapter(SourceAdapter):
    slug = "ce"

    def __init__(
        self,
        *,
        definition: CEReleaseDefinition | None = None,
        client: httpx.Client | None = None,
    ) -> None:
        self.definition = definition or _load_definition(_DEFINITION_PATH)
        self.client = client or build_bls_client()

    def available_years(self) -> list[int]:
        self._landing_html()
        return [self.definition.year]

    def discover(self, year: int | None = None) -> DiscoveredRelease:
        selected_year = year or self.definition.year
        if selected_year != self.definition.year:
            raise ValueError(f"CE release definition unavailable for {selected_year}")
        artifacts = _page_artifacts(
            self._landing_html(),
            str(self.definition.landing_page),
        )
        artifacts.extend(
            DiscoveredArtifact(
                role=item.role,
                url=item.url,
                filename=item.filename,
                documentation=item.documentation,
                link_text=item.filename,
            )
            for item in self.definition.static_artifacts
        )
        roles = {_role_value(item.role) for item in artifacts}
        missing = set(self.definition.required_artifacts) - roles
        if missing:
            raise ValueError(f"CE landing page is missing required roles: {sorted(missing)}")
        weight = self.definition.weight_contract
        benchmark = self.definition.benchmark
        return DiscoveredRelease(
            survey=self.slug,
            year=selected_year,
            landing_page=self.definition.landing_page,
            artifacts=artifacts,
            source_metadata={
                "record_unit": "consumer unit interview-quarter and diary-week records",
                "universe": "U.S. civilian noninstitutional consumer units",
                "calendar_year": selected_year,
                "weight": weight.full_sample,
                "replicate_weights": weight.replicate_count,
                "variance_factor": weight.variance_factor,
                "benchmark_ucc": benchmark.ucc,
                "benchmark_source": benchmark.source_code,
                "adjacent_archive_required": True,
                "interview_diary_integration_required": True,
            },
        )

    def validate_release(
        self,
        run_root: Path,
        release: DiscoveredRelease,
        artifacts: list[StoredArtifact],
    ) -> ValidationResult:
        roles = {_role_value(item.role) for item in artifacts}
        checks: dict[str, bool] = {
            "ce_required_artifacts_present": set(self.definition.required_artifacts) <= roles,
            "ce_release_year_matches": release.year == self.definition.year,
        }
        notes = [
            "record_unit=consumer unit interview-quarter and diary-week records",
            "interview_calendar_year=five quarter files with adjacent archive",
            "diary_calendar_year=four collection-quarter files",
            "integration=official year-specific UCC source assignment required",
        ]

        family_frames: list[pd.DataFrame] = []
        expenditure_frames: list[pd.DataFrame] = []
        selected_paths: list[str] = []
        weights = self.definition.weight_contract
        replicate_names = [
            f"{weights.replicate_prefix}{number:02d}"
            for number in range(1, weights.replicate_count + 1)
        ]
        family_columns = [
            "NEWID",
            weights.full_sample,
            "QINTRVMO",
            "QINTRVYR",
            *replicate_names,
        ]
        expenditure_columns = ["NEWID", "COST", "UCC", "REF_YR"]

        for member in self.definition.members:
            source = _extracted_member(run_root, artifacts, member.role, member.path)
            selected_paths.append(member.path)
            header = _csv_header(source)
            key = _check_key(member.path)
            checks[f"ce_{key}_columns"] = len(header) == member.columns
            if member.family == "interview_family":
                frame = pd.read_csv(source, usecols=family_columns, low_memory=False)
                family_frames.append(frame)
                checks[f"ce_{key}_rows"] = len(frame) == member.rows
                checks[f"ce_{key}_required_columns"] = set(family_columns) <= set(header)
            elif member.family == "interview_expenditures":
                frame = pd.read_csv(source, usecols=expenditure_columns, low_memory=False)
                expenditure_frames.append(frame)
                checks[f"ce_{key}_rows"] = len(frame) == member.rows
                checks[f"ce_{key}_required_columns"] = set(expenditure_columns) <= set(header)
            else:
                checks[f"ce_{key}_rows"] = _csv_data_rows(source) == member.rows
        checks["ce_interview_overlap_copy_excluded"] = not any(
            "241x" in path.lower() for path in selected_paths
        )
        checks["ce_member_contract_complete"] = len(selected_paths) == len(self.definition.members)

        family = pd.concat(family_frames, ignore_index=True)
        expenditures = pd.concat(expenditure_frames, ignore_index=True)
        checks["ce_interview_newid_unique"] = bool(family["NEWID"].is_unique)
        full_weight = pd.to_numeric(family[weights.full_sample], errors="coerce")
        checks["ce_full_weight_positive"] = bool(
            full_weight.notna().all() and (full_weight > 0).all()
        )
        replicate_frame = family[replicate_names].apply(pd.to_numeric, errors="coerce")
        checks["ce_replicate_count"] = len(replicate_names) == weights.replicate_count
        checks["ce_replicates_positive_or_missing"] = bool(
            (replicate_frame.isna() | (replicate_frame > 0)).all().all()
        )
        checks["ce_replicates_have_balanced_missingness"] = bool(
            replicate_frame.isna().any().all() and replicate_frame.notna().any().all()
        )

        benchmark = self.definition.benchmark
        ucc = pd.to_numeric(expenditures["UCC"], errors="coerce")
        ref_year = pd.to_numeric(expenditures["REF_YR"], errors="coerce")
        cost = pd.to_numeric(expenditures["COST"], errors="coerce")
        selected = expenditures.loc[
            (ucc == int(benchmark.ucc)) & (ref_year == benchmark.reference_year),
            ["NEWID"],
        ].copy()
        selected["COST"] = cost.loc[selected.index]
        by_newid = selected.groupby("NEWID", as_index=False)["COST"].sum()
        analysis = family.merge(by_newid, on="NEWID", how="left")
        analysis["COST"] = analysis["COST"].fillna(0.0)
        interview_month = pd.to_numeric(analysis["QINTRVMO"], errors="coerce")
        interview_year = pd.to_numeric(analysis["QINTRVYR"], errors="coerce")
        calendar_factor = np.where(
            (interview_year == benchmark.reference_year) & interview_month.isin([1, 2, 3]),
            (interview_month - 1.0) / 3.0,
            np.where(
                interview_year == benchmark.reference_year + 1,
                (4.0 - interview_month) / 3.0,
                1.0,
            ),
        )
        full_numerator = float((analysis["COST"] * full_weight).sum())
        full_denominator = float((calendar_factor * full_weight / 4.0).sum())
        observed_mean = full_numerator / full_denominator
        replicate_estimates = []
        for name in replicate_names:
            replicate_weight = pd.to_numeric(analysis[name], errors="coerce")
            numerator = float((analysis["COST"] * replicate_weight).sum())
            denominator = float((calendar_factor * replicate_weight / 4.0).sum())
            replicate_estimates.append(numerator / denominator)
        observed_se = math.sqrt(
            weights.variance_factor
            * sum((estimate - observed_mean) ** 2 for estimate in replicate_estimates)
        )
        observed_population_thousands = full_denominator / 1000.0
        observed_reporters = int((analysis["COST"] > 0).sum())
        checks["ce_benchmark_mean"] = (
            abs(observed_mean - benchmark.expected_mean) <= benchmark.rounding_tolerance
        )
        checks["ce_benchmark_standard_error"] = (
            abs(observed_se - benchmark.expected_standard_error) <= benchmark.rounding_tolerance
        )
        checks["ce_benchmark_population"] = (
            abs(observed_population_thousands - benchmark.expected_population_thousands)
            <= benchmark.population_rounding_tolerance
        )
        checks["ce_benchmark_positive_reporters"] = (
            observed_reporters == benchmark.expected_positive_reporters
        )
        checks["ce_benchmark_reference_year_domain"] = set(interview_year.dropna().astype(int)) == {
            benchmark.reference_year,
            benchmark.reference_year + 1,
        }

        checks.update(
            _validate_workbooks_and_code(
                run_root,
                artifacts,
                self.definition,
            )
        )
        notes.extend(
            [
                f"benchmark_ucc={benchmark.ucc}",
                f"benchmark_observed_mean={observed_mean:.8f}",
                f"benchmark_expected_mean={benchmark.expected_mean:.2f}",
                f"benchmark_observed_se={observed_se:.8f}",
                f"benchmark_expected_se={benchmark.expected_standard_error:.2f}",
                f"weighted_consumer_units_thousands={observed_population_thousands:.6f}",
                f"positive_reporters={observed_reporters}",
                "uncertainty=44 balanced repeated replication half-samples",
            ]
        )
        return ValidationResult(
            passed=all(checks.values()),
            checks=checks,
            notes=notes,
        )

    def normalize_release(
        self,
        run_root: Path,
        release: DiscoveredRelease,
        artifacts: list[StoredArtifact],
    ) -> list[Path]:
        del release
        normalized = run_root / "normalized"
        normalized.mkdir(parents=True, exist_ok=True)
        outputs: list[Path] = []
        connection = duckdb.connect()
        try:
            families = sorted({member.family for member in self.definition.members})
            for family in families:
                members = [member for member in self.definition.members if member.family == family]
                sources = [
                    _extracted_member(run_root, artifacts, member.role, member.path)
                    for member in members
                ]
                output = normalized / f"{family}.parquet"
                _stack_csv_to_parquet(connection, sources, output)
                outputs.append(output)
        finally:
            connection.close()
        return outputs

    def download_client(self) -> httpx.Client:
        return self.client

    def close(self) -> None:
        self.client.close()

    def _landing_html(self) -> str:
        response = require_bls_response(self.client.get(str(self.definition.landing_page)))
        return response.text


def _load_definition(path: Path) -> CEReleaseDefinition:
    with path.open() as handle:
        return CEReleaseDefinition.model_validate(yaml.safe_load(handle))


def _page_artifacts(html: str, landing_page: str) -> list[DiscoveredArtifact]:
    targets = {
        "/cex/pumd/data/csv/intrvw23.zip": (INTERVIEW_PRIOR, False),
        "/cex/pumd/data/csv/intrvw24.zip": (INTERVIEW_CURRENT, False),
        "/cex/pumd/data/csv/diary24.zip": (DIARY_CURRENT, False),
        "/cex/pumd/ce-pumd-interview-diary-dictionary.xlsx": (DICTIONARY, False),
        "/cex/pumd/errata/errata-2023.htm": (PRIOR_ERRATA, True),
    }
    found: dict[str, DiscoveredArtifact] = {}
    tree = HTMLParser(html)
    for node in tree.css("a"):
        href = node.attributes.get("href")
        if not href:
            continue
        url = urljoin(landing_page, href)
        path = urlparse(url).path.lower()
        if path not in targets:
            continue
        role, documentation = targets[path]
        filename = Path(unquote(urlparse(url).path)).name
        found[role] = DiscoveredArtifact(
            role=role,
            url=_url_adapter.validate_python(url),
            filename=filename,
            documentation=documentation,
            link_text=node.text(strip=True) or filename,
        )
    missing = set(role for role, _ in targets.values()) - set(found)
    if missing:
        raise ValueError(f"CE landing page is missing exact CSV/document links: {sorted(missing)}")
    return [found[role] for role, _ in targets.values()]


def _validate_workbooks_and_code(
    run_root: Path,
    artifacts: list[StoredArtifact],
    definition: CEReleaseDefinition,
) -> dict[str, bool]:
    checks: dict[str, bool] = {}
    benchmark_path = _artifact_path(run_root, artifacts, BENCHMARK_WORKBOOK)
    integration_path = _artifact_path(run_root, artifacts, INTEGRATION_WORKBOOK)
    dictionary_path = _artifact_path(run_root, artifacts, DICTIONARY)
    sample_paths = _extracted_for_role(run_root, artifacts, SAMPLE_R)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        benchmark_sheet = load_workbook(benchmark_path, data_only=True, read_only=True)[
            definition.benchmark.workbook_sheet
        ]
        benchmark_rows = list(benchmark_sheet.values)
        label_index = next(
            (
                index
                for index, row in enumerate(benchmark_rows)
                if row and row[0] == definition.benchmark.workbook_label
            ),
            None,
        )
        checks["ce_official_benchmark_label_present"] = label_index is not None
        if label_index is not None:
            mean_value = float(str(benchmark_rows[label_index + 1][1]))
            se_value = float(str(benchmark_rows[label_index + 3][1]))
            checks["ce_official_benchmark_mean_contract"] = (
                mean_value == definition.benchmark.expected_mean
            )
            checks["ce_official_benchmark_se_contract"] = (
                se_value == definition.benchmark.expected_standard_error
            )

        integration_rows = list(
            load_workbook(integration_path, data_only=True, read_only=True)[
                "AllYears_IntStub"
            ].values
        )
        header = {str(value): index for index, value in enumerate(integration_rows[2])}
        integration_match = any(
            str(row[header["UCC"]]).strip() == definition.benchmark.ucc
            and str(row[header["y24"]]).strip() == definition.benchmark.source_code
            and definition.benchmark.label.lower() in str(row[header["Description"]]).lower()
            for row in integration_rows[3:]
        )
        checks["ce_official_integration_source_contract"] = integration_match

        dictionary_rows = list(
            load_workbook(dictionary_path, data_only=True, read_only=True)[
                definition.dictionary.sheet
            ].values
        )
        dictionary_header = {str(value): index for index, value in enumerate(dictionary_rows[0])}
        variable_index = dictionary_header.get("Variable Name")
        variables = (
            {str(row[variable_index]).strip() for row in dictionary_rows[1:]}
            if variable_index is not None
            else set()
        )
        checks["ce_dictionary_variable_contract"] = (
            set(definition.dictionary.required_variables) <= variables
        )

    r_files = [path for path in sample_paths if path.suffix.lower() == ".r"]
    checks["ce_sample_r_present"] = len(r_files) == 1
    if r_files:
        sample_text = r_files[0].read_text(errors="replace")
        checks["ce_sample_r_calendar_formula_present"] = all(
            token in sample_text for token in ("FINLWT21", "QINTRVMO", "QINTRVYR", "popwt")
        )
    return checks


def _stack_csv_to_parquet(
    connection: duckdb.DuckDBPyConnection,
    sources: list[Path],
    output: Path,
) -> None:
    source_sql = ", ".join(_sql_literal(str(path)) for path in sources)
    query = f"""
        COPY (
            SELECT * EXCLUDE (filename),
                   regexp_extract(filename, '[^/]+$') AS SOURCE_FILE
            FROM read_csv_auto(
                [{source_sql}],
                header = true,
                union_by_name = true,
                filename = true,
                sample_size = -1
            )
        ) TO {_sql_literal(str(output))}
        (FORMAT PARQUET, COMPRESSION ZSTD)
    """
    connection.execute(query)


def _sql_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def _artifact_path(
    run_root: Path,
    artifacts: list[StoredArtifact],
    role: str,
) -> Path:
    matches = [item for item in artifacts if _role_value(item.role) == role]
    if len(matches) != 1:
        raise ValueError(f"Expected one CE artifact for {role}, found {len(matches)}")
    return run_root / matches[0].relative_path


def _extracted_for_role(
    run_root: Path,
    artifacts: list[StoredArtifact],
    role: str,
) -> list[Path]:
    matches = [item for item in artifacts if _role_value(item.role) == role]
    if len(matches) != 1:
        raise ValueError(f"Expected one CE artifact for {role}, found {len(matches)}")
    return [run_root / relative for relative in matches[0].extracted_files]


def _extracted_member(
    run_root: Path,
    artifacts: list[StoredArtifact],
    role: str,
    member_path: str,
) -> Path:
    suffix = member_path.replace("\\", "/").lower()
    matches = [
        path
        for path in _extracted_for_role(run_root, artifacts, role)
        if path.as_posix().lower().endswith(suffix)
    ]
    if len(matches) != 1:
        raise ValueError(f"Expected one CE member {member_path} for {role}, found {len(matches)}")
    return matches[0]


def _csv_header(path: Path) -> list[str]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return next(csv.reader(handle))


def _csv_data_rows(path: Path) -> int:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        next(reader)
        return sum(1 for _ in reader)


def _check_key(path: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", path.lower()).strip("_")


def _role_value(role: object) -> str:
    value = getattr(role, "value", role)
    return str(value)
