from __future__ import annotations

import csv
import json
import math
import warnings
from collections.abc import Iterable
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

import httpx
import numpy as np
import pandas as pd
import pyarrow as pa
import pyarrow.csv as pacsv
import pyarrow.parquet as pq
import yaml
from openpyxl import load_workbook
from pydantic import AnyHttpUrl, BaseModel, ConfigDict
from selectolax.parser import HTMLParser

from microdata_lab.adapters.base import SourceAdapter
from microdata_lab.models import (
    ArtifactRole,
    DiscoveredArtifact,
    DiscoveredRelease,
    StoredArtifact,
    ValidationResult,
)

PRIMARY_DATA = "primary_data"
PRIMARY_SCHEMA = "primary_schema"
PRIMARY_VALIDATION = "primary_validation"
REPLICATE_WEIGHTS = "replicate_weights"
REPLICATE_SCHEMA = "replicate_schema"
REPLICATE_VALIDATION = "replicate_validation"
CODEBOOK = "codebook"
RELEASE_NOTES = "release_notes"
INPUT_EXAMPLE = "input_example"
USERS_GUIDE = "users_guide"

_DEFINITION_PATH = Path(__file__).resolve().parents[3] / "config" / "sipp" / "2025.yaml"


class SIPPBenchmarkDefinition(BaseModel):
    variable: str
    expected_nonmissing: int
    expected_mean: float
    absolute_tolerance: float
    validation_workbook_mean: float


class SIPPWeightContract(BaseModel):
    primary: str
    replicate_full_sample: str
    replicate_prefix: str
    variance_replicates: int
    fay_adjustment: float
    expected_primary_mean: float
    expected_replicate_mean: float
    mean_tolerance: float


class SIPPReleaseDefinition(BaseModel):
    model_config = ConfigDict(arbitrary_types_allowed=True)

    year: int
    landing_page: AnyHttpUrl
    users_guide_url: AnyHttpUrl
    required_roles: list[str]
    expected_primary_rows: int
    expected_primary_columns: int
    expected_replicate_rows: int
    expected_replicate_columns: int
    expected_weight_columns: int
    expected_positive_weight_rows: int
    expected_panel_min: int
    expected_months: list[int]
    benchmark: SIPPBenchmarkDefinition
    weight_contract: SIPPWeightContract


class SIPPAdapter(SourceAdapter):
    slug = "sipp"

    def __init__(
        self,
        *,
        definition: SIPPReleaseDefinition | None = None,
        client: httpx.Client | None = None,
    ) -> None:
        self.definition = definition or _load_definition(_DEFINITION_PATH)
        self._owns_client = client is None
        self.client = client or httpx.Client(
            follow_redirects=True,
            timeout=httpx.Timeout(120.0, connect=30.0),
            headers={"User-Agent": "microdata-lab/0.1 (public microdata research)"},
        )

    def close(self) -> None:
        if self._owns_client:
            self.client.close()

    def available_years(self) -> list[int]:
        links = self._release_links()
        return [self.definition.year] if self._has_complete_release(links) else []

    def discover(self, year: int | None = None) -> DiscoveredRelease:
        requested_year = year or self.definition.year
        if requested_year != self.definition.year:
            raise ValueError(
                f"SIPP {requested_year} has no pinned benchmark definition; "
                f"available={self.definition.year}"
            )
        links = self._release_links()
        expected = _expected_filenames(requested_year)
        artifacts: list[DiscoveredArtifact] = []
        for role in self.definition.required_roles:
            filename = expected[role]
            matches = [url for url in links if Path(urlparse(url).path).name == filename]
            if len(matches) != 1:
                raise ValueError(
                    f"SIPP {requested_year} requires exactly one {role} artifact named "
                    f"{filename}; found={len(matches)}"
                )
            artifacts.append(_artifact(role, matches[0]))
        artifacts.append(_artifact(USERS_GUIDE, str(self.definition.users_guide_url)))
        return DiscoveredRelease(
            survey=self.slug,
            year=requested_year,
            landing_page=self.definition.landing_page,
            artifacts=artifacts,
            source_metadata={
                "record_unit": "person-month",
                "universe": "civilian noninstitutionalized population of the United States",
                "reference_period": f"January through December {requested_year - 1}",
                "primary_weight": self.definition.weight_contract.primary,
                "replicate_weight_full_sample": (
                    self.definition.weight_contract.replicate_full_sample
                ),
                "variance_replicates": self.definition.weight_contract.variance_replicates,
                "fay_adjustment": self.definition.weight_contract.fay_adjustment,
                "merge_keys": ["SSUID", "PNUM", "MONTHCODE"],
                "benchmark": self.definition.benchmark.model_dump(),
                "release_status": "official public-use version",
            },
        )

    def validate_release(
        self,
        run_root: Path,
        release: DiscoveredRelease,
        artifacts: list[StoredArtifact],
    ) -> ValidationResult:
        checks: dict[str, bool] = {}
        notes: list[str] = []
        roles = {_role_value(artifact.role) for artifact in artifacts}
        required = set(self.definition.required_roles) | {USERS_GUIDE}
        checks["sipp_required_artifacts_present"] = required <= roles

        primary_csv = _single_extracted(run_root, artifacts, PRIMARY_DATA, ".csv")
        replicate_csv = _single_extracted(run_root, artifacts, REPLICATE_WEIGHTS, ".csv")
        primary_schema_path = _single_stored(run_root, artifacts, PRIMARY_SCHEMA)
        replicate_schema_path = _single_stored(run_root, artifacts, REPLICATE_SCHEMA)
        primary_validation_path = _single_stored(run_root, artifacts, PRIMARY_VALIDATION)
        replicate_validation_path = _single_stored(run_root, artifacts, REPLICATE_VALIDATION)

        primary_schema = _load_schema(primary_schema_path)
        replicate_schema = _load_schema(replicate_schema_path)
        primary_names = [item["name"] for item in primary_schema]
        replicate_names = [item["name"] for item in replicate_schema]
        replicate_weight_names = [
            name
            for name in replicate_names
            if name.startswith(self.definition.weight_contract.replicate_prefix)
        ]
        replicate_header = [name.upper() for name in _pipe_header(replicate_csv)]
        checks.update(
            {
                "sipp_primary_schema_column_count": len(primary_names)
                == self.definition.expected_primary_columns,
                "sipp_replicate_schema_column_count": len(replicate_names)
                == self.definition.expected_replicate_columns,
                "sipp_replicate_weight_count": len(replicate_weight_names)
                == self.definition.expected_weight_columns,
                "sipp_replicate_weight_sequence": replicate_weight_names
                == [
                    f"{self.definition.weight_contract.replicate_prefix}{index}"
                    for index in range(self.definition.weight_contract.variance_replicates + 1)
                ],
                "sipp_primary_header_matches_schema": _pipe_header(primary_csv) == primary_names,
                "sipp_replicate_header_matches_schema": replicate_header
                == [name.upper() for name in replicate_names],
            }
        )

        primary_stats = _workbook_stats(primary_validation_path)
        replicate_stats = _workbook_stats(replicate_validation_path)
        benchmark_name = self.definition.benchmark.variable.upper()
        primary_weight_name = self.definition.weight_contract.primary.upper()
        replicate_full_name = self.definition.weight_contract.replicate_full_sample.upper()
        checks.update(
            {
                "sipp_validation_workbook_primary_rows": _stat_int(primary_stats, "SPANEL", "N")
                == self.definition.expected_primary_rows,
                "sipp_validation_workbook_replicate_rows": _stat_int(replicate_stats, "SPANEL", "N")
                == self.definition.expected_replicate_rows,
                "sipp_validation_workbook_benchmark_nonmissing": _stat_int(
                    primary_stats, benchmark_name, "N"
                )
                == self.definition.benchmark.expected_nonmissing,
                "sipp_validation_workbook_benchmark_mean": math.isclose(
                    _stat_float(primary_stats, benchmark_name, "MEAN"),
                    self.definition.benchmark.validation_workbook_mean,
                    abs_tol=self.definition.benchmark.absolute_tolerance,
                ),
            }
        )

        key_columns = ["SSUID", "PNUM", "MONTHCODE"]
        primary_columns = [
            *key_columns,
            "SPANEL",
            "SWAVE",
            "RIN_UNIV",
            self.definition.weight_contract.primary,
            self.definition.benchmark.variable,
        ]
        primary_rows = 0
        primary_keys: set[tuple[str, int, int]] = set()
        positive_weight_by_key: dict[tuple[str, int, int], float] = {}
        primary_duplicates = False
        primary_weight_sum = 0.0
        primary_weight_nonnegative = True
        benchmark_sum = 0.0
        benchmark_nonmissing = 0
        panels: set[int] = set()
        waves: set[int] = set()
        months: set[int] = set()
        universe_codes: set[int] = set()

        for chunk in pd.read_csv(
            primary_csv,
            sep="|",
            usecols=primary_columns,
            chunksize=5_000,
            low_memory=False,
        ):
            primary_rows += len(chunk)
            weights = pd.to_numeric(chunk[self.definition.weight_contract.primary], errors="coerce")
            values = pd.to_numeric(chunk[self.definition.benchmark.variable], errors="coerce")
            primary_weight_nonnegative &= bool(weights.notna().all() and (weights >= 0).all())
            primary_weight_sum += float(weights.sum())
            valid_values = values.dropna()
            benchmark_sum += float(valid_values.sum())
            benchmark_nonmissing += len(valid_values)
            panels.update(int(value) for value in chunk["SPANEL"].dropna().unique())
            waves.update(int(value) for value in chunk["SWAVE"].dropna().unique())
            months.update(int(value) for value in chunk["MONTHCODE"].dropna().unique())
            universe_codes.update(int(value) for value in chunk["RIN_UNIV"].dropna().unique())

            for ssuid, pnum, month, weight in zip(
                chunk["SSUID"].astype(str),
                chunk["PNUM"],
                chunk["MONTHCODE"],
                weights,
                strict=True,
            ):
                key = (ssuid, int(pnum), int(month))
                if key in primary_keys:
                    primary_duplicates = True
                primary_keys.add(key)
                if float(weight) > 0:
                    positive_weight_by_key[key] = float(weight)

        benchmark_mean = benchmark_sum / benchmark_nonmissing if benchmark_nonmissing else math.nan
        primary_weight_mean = primary_weight_sum / primary_rows if primary_rows else math.nan

        replicate_rows = 0
        replicate_keys: set[tuple[str, int, int]] = set()
        replicate_duplicates = False
        replicate_weights_valid = True
        replicate_full_sum = 0.0
        replicate_full_matches = True
        replicate_read_columns = key_columns + replicate_weight_names
        for chunk in pd.read_csv(
            replicate_csv,
            sep="|",
            names=replicate_names,
            header=0,
            usecols=replicate_read_columns,
            chunksize=2_000,
            low_memory=False,
        ):
            replicate_rows += len(chunk)
            weights = chunk[replicate_weight_names].apply(pd.to_numeric, errors="coerce")
            weight_values = weights.to_numpy(dtype=float)
            replicate_weights_valid &= bool(
                np.isfinite(weight_values).all() and (weight_values > 0).all()
            )
            full_weights = weights[self.definition.weight_contract.replicate_full_sample]
            replicate_full_sum += float(full_weights.sum())
            for ssuid, pnum, month, full_weight in zip(
                chunk["SSUID"].astype(str),
                chunk["PNUM"],
                chunk["MONTHCODE"],
                full_weights,
                strict=True,
            ):
                key = (ssuid, int(pnum), int(month))
                if key in replicate_keys:
                    replicate_duplicates = True
                replicate_keys.add(key)
                primary_weight = positive_weight_by_key.get(key)
                if primary_weight is None or not math.isclose(
                    primary_weight, float(full_weight), abs_tol=1e-6
                ):
                    replicate_full_matches = False

        replicate_full_mean = replicate_full_sum / replicate_rows if replicate_rows else math.nan
        checks.update(
            {
                "sipp_primary_row_count": primary_rows == self.definition.expected_primary_rows,
                "sipp_primary_person_month_keys_unique": not primary_duplicates
                and len(primary_keys) == primary_rows,
                "sipp_primary_weights_nonnegative": primary_weight_nonnegative,
                "sipp_positive_weight_row_count": len(positive_weight_by_key)
                == self.definition.expected_positive_weight_rows,
                "sipp_primary_weight_mean": math.isclose(
                    primary_weight_mean,
                    self.definition.weight_contract.expected_primary_mean,
                    abs_tol=self.definition.weight_contract.mean_tolerance,
                ),
                "sipp_panel_range": panels
                == set(range(self.definition.expected_panel_min, release.year + 1)),
                "sipp_wave_domain": waves
                == set(range(1, release.year - self.definition.expected_panel_min + 2)),
                "sipp_month_domain": months == set(self.definition.expected_months),
                "sipp_universe_code_domain": universe_codes <= {1, 2},
                "sipp_benchmark_nonmissing": benchmark_nonmissing
                == self.definition.benchmark.expected_nonmissing,
                "sipp_official_validation_benchmark": math.isclose(
                    benchmark_mean,
                    self.definition.benchmark.expected_mean,
                    abs_tol=self.definition.benchmark.absolute_tolerance,
                ),
                "sipp_replicate_row_count": replicate_rows
                == self.definition.expected_replicate_rows,
                "sipp_replicate_person_month_keys_unique": not replicate_duplicates
                and len(replicate_keys) == replicate_rows,
                "sipp_replicate_keys_match_positive_primary": replicate_keys
                == set(positive_weight_by_key),
                "sipp_replicate_weights_positive": replicate_weights_valid,
                "sipp_repwgt0_matches_primary_weight": replicate_full_matches
                and math.isclose(replicate_full_sum, primary_weight_sum, abs_tol=1e-3),
                "sipp_replicate_weight_mean": math.isclose(
                    replicate_full_mean,
                    self.definition.weight_contract.expected_replicate_mean,
                    abs_tol=self.definition.weight_contract.mean_tolerance,
                ),
                "sipp_workbook_primary_weight_mean": math.isclose(
                    _stat_float(primary_stats, primary_weight_name, "MEAN"),
                    self.definition.weight_contract.expected_primary_mean,
                    abs_tol=self.definition.weight_contract.mean_tolerance,
                ),
                "sipp_workbook_replicate_weight_mean": math.isclose(
                    _stat_float(replicate_stats, replicate_full_name, "MEAN"),
                    self.definition.weight_contract.expected_replicate_mean,
                    abs_tol=self.definition.weight_contract.mean_tolerance,
                ),
            }
        )
        notes.extend(
            [
                f"primary_rows={primary_rows}",
                f"primary_columns={len(primary_names)}",
                f"replicate_rows={replicate_rows}",
                f"replicate_columns={len(replicate_names)}",
                f"positive_weight_rows={len(positive_weight_by_key)}",
                f"benchmark_variable={self.definition.benchmark.variable}",
                f"benchmark_observed={benchmark_mean:.8f}",
                f"benchmark_expected={self.definition.benchmark.expected_mean:.8f}",
                f"primary_weight={self.definition.weight_contract.primary}",
                "replicate_weights=REPWGT1-REPWGT240; REPWGT0 is the full-sample weight",
                "variance=1/(240*0.5^2) times sum of squared replicate deviations",
                "record_unit=person-month; unique person=SSUID+PNUM",
                "imputation=variable-level status flags documented in the Census dictionary",
                "limitation=2025 release reports lower-than-average unit response",
            ]
        )
        return ValidationResult(passed=all(checks.values()), checks=checks, notes=notes)

    def normalize_release(
        self,
        run_root: Path,
        release: DiscoveredRelease,
        artifacts: list[StoredArtifact],
    ) -> list[Path]:
        primary_csv = _single_extracted(run_root, artifacts, PRIMARY_DATA, ".csv")
        replicate_csv = _single_extracted(run_root, artifacts, REPLICATE_WEIGHTS, ".csv")
        primary_schema = _single_stored(run_root, artifacts, PRIMARY_SCHEMA)
        replicate_schema = _single_stored(run_root, artifacts, REPLICATE_SCHEMA)
        normalized = run_root / "normalized"
        normalized.mkdir(parents=True, exist_ok=True)
        primary_output = normalized / "person_month.parquet"
        replicate_output = normalized / "replicate_weights.parquet"
        _pipe_csv_to_parquet(primary_csv, primary_schema, primary_output)
        _pipe_csv_to_parquet(replicate_csv, replicate_schema, replicate_output)
        return [primary_output, replicate_output]

    def _release_links(self) -> set[str]:
        response = self.client.get(str(self.definition.landing_page))
        response.raise_for_status()
        normalized = response.text.lower()
        if "page not found" in normalized or "access denied" in normalized:
            raise ValueError("SIPP release page returned denial or not-found content")
        page = HTMLParser(response.text)
        return {
            urljoin(str(self.definition.landing_page), anchor.attributes["href"])
            for anchor in page.css("a[href]")
            if anchor.attributes.get("href")
        }

    def _has_complete_release(self, links: set[str]) -> bool:
        expected = _expected_filenames(self.definition.year)
        filenames = {Path(urlparse(url).path).name for url in links}
        return all(expected[role] in filenames for role in self.definition.required_roles)


def _load_definition(path: Path) -> SIPPReleaseDefinition:
    return SIPPReleaseDefinition.model_validate(yaml.safe_load(path.read_text()))


def _expected_filenames(year: int) -> dict[str, str]:
    return {
        PRIMARY_DATA: f"pu{year}_csv.zip",
        PRIMARY_SCHEMA: f"pu{year}_schema.json",
        PRIMARY_VALIDATION: f"pu{year}_validate.xlsx",
        REPLICATE_WEIGHTS: f"rw{year}_csv.zip",
        REPLICATE_SCHEMA: f"rw{year}_schema.json",
        REPLICATE_VALIDATION: f"rw{year}_validate.xlsx",
        CODEBOOK: f"{year}_SIPP_Data_Dictionary.pdf",
        RELEASE_NOTES: f"{year}_SIPP_Release_Notes.pdf",
        INPUT_EXAMPLE: f"{year}_sipp_python_input_example.py",
    }


def _artifact(role: str, url: str) -> DiscoveredArtifact:
    filename = Path(urlparse(url).path).name
    return DiscoveredArtifact(
        role=role,
        url=AnyHttpUrl(url),
        link_text=filename,
        filename=filename,
        documentation=role in {CODEBOOK, RELEASE_NOTES, USERS_GUIDE},
    )


def _role_value(role: ArtifactRole | str) -> str:
    return role.value if isinstance(role, ArtifactRole) else role


def _single_stored(run_root: Path, artifacts: Iterable[StoredArtifact], role: str) -> Path:
    matches = [
        run_root / artifact.relative_path
        for artifact in artifacts
        if _role_value(artifact.role) == role
    ]
    if len(matches) != 1:
        raise ValueError(f"SIPP requires one stored {role} artifact; found={len(matches)}")
    return matches[0]


def _single_extracted(
    run_root: Path,
    artifacts: Iterable[StoredArtifact],
    role: str,
    suffix: str,
) -> Path:
    matches: list[Path] = []
    for artifact in artifacts:
        if _role_value(artifact.role) != role:
            continue
        matches.extend(
            run_root / relative
            for relative in artifact.extracted_files
            if relative.lower().endswith(suffix)
        )
    if len(matches) != 1:
        raise ValueError(
            f"SIPP requires one extracted {suffix} file for {role}; found={len(matches)}"
        )
    return matches[0]


def _load_schema(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text())
    if not isinstance(data, list) or not data:
        raise ValueError(f"invalid SIPP schema: {path.name}")
    for item in data:
        if not isinstance(item, dict) or not {"name", "dtype"} <= item.keys():
            raise ValueError(f"invalid SIPP schema entry in {path.name}")
    return data


def _pipe_header(path: Path) -> list[str]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return next(csv.reader(handle, delimiter="|"))


def _workbook_stats(path: Path) -> dict[str, dict[str, Any]]:
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Cannot parse header or footer")
        workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        raise ValueError(f"SIPP validation workbook has no active sheet: {path.name}")
    rows = sheet.iter_rows(values_only=True)
    header = [str(value).strip().upper() for value in next(rows)]
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        values = dict(zip(header, row, strict=False))
        variable = values.get("VARIABLE")
        if variable:
            result[str(variable).upper()] = values
    workbook.close()
    return result


def _stat_int(stats: dict[str, dict[str, Any]], variable: str, field: str) -> int:
    return int(stats[variable.upper()][field.upper()])


def _stat_float(stats: dict[str, dict[str, Any]], variable: str, field: str) -> float:
    return float(stats[variable.upper()][field.upper()])


def _pipe_csv_to_parquet(csv_path: Path, schema_path: Path, output: Path) -> None:
    schema = _load_schema(schema_path)
    type_map: dict[str, pa.DataType] = {
        "integer": pa.int64(),
        "float": pa.float64(),
        "string": pa.string(),
    }
    column_types = {str(item["name"]): type_map[str(item["dtype"])] for item in schema}
    reader = pacsv.open_csv(
        csv_path,
        read_options=pacsv.ReadOptions(
            block_size=16 * 1024 * 1024,
            column_names=list(column_types),
            skip_rows=1,
            use_threads=True,
        ),
        parse_options=pacsv.ParseOptions(delimiter="|"),
        convert_options=pacsv.ConvertOptions(
            column_types=column_types,
            strings_can_be_null=True,
        ),
    )
    with pq.ParquetWriter(output, reader.schema, compression="zstd") as writer:
        for batch in reader:
            writer.write_batch(batch)
