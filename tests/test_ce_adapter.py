from __future__ import annotations

from pathlib import Path

import httpx
import pandas as pd
import pyarrow.parquet as pq
from openpyxl import Workbook
from pydantic import AnyHttpUrl, TypeAdapter

from microdata_lab.adapters.ce import (
    BENCHMARK_WORKBOOK,
    DIARY_CURRENT,
    DICTIONARY,
    GETTING_STARTED,
    INTEGRATION_WORKBOOK,
    INTERVIEW_CURRENT,
    INTERVIEW_PRIOR,
    PRIOR_ERRATA,
    SAMPLE_R,
    CEAdapter,
    CEBenchmark,
    CEDictionaryContract,
    CEMemberContract,
    CEReleaseDefinition,
    CEStaticArtifact,
    CEWeightContract,
)
from microdata_lab.models import StoredArtifact

_url_adapter = TypeAdapter(AnyHttpUrl)
LANDING = "https://www.bls.gov/cex/pumd_data.htm"
PAGE_ROLES = {
    INTERVIEW_PRIOR,
    INTERVIEW_CURRENT,
    DIARY_CURRENT,
    DICTIONARY,
    PRIOR_ERRATA,
}
STATIC_ROLES = {
    GETTING_STARTED,
    INTEGRATION_WORKBOOK,
    BENCHMARK_WORKBOOK,
    SAMPLE_R,
}


def _definition() -> CEReleaseDefinition:
    static = [
        CEStaticArtifact(
            role=role,
            url=_url_adapter.validate_python(f"https://www.bls.gov/fixture/{role}"),
            filename=f"{role}.{'zip' if role == SAMPLE_R else 'xlsx'}",
            documentation=role == GETTING_STARTED,
        )
        for role in sorted(STATIC_ROLES)
    ]
    return CEReleaseDefinition(
        year=2024,
        landing_page=_url_adapter.validate_python(LANDING),
        required_artifacts=sorted(PAGE_ROLES | STATIC_ROLES),
        static_artifacts=static,
        members=[
            CEMemberContract(
                role=INTERVIEW_PRIOR,
                path="intrvw23/fmli241.csv",
                family="interview_family",
                rows=1,
                columns=9,
                quarter="241",
            ),
            CEMemberContract(
                role=INTERVIEW_PRIOR,
                path="intrvw23/mtbi241.csv",
                family="interview_expenditures",
                rows=1,
                columns=4,
                quarter="241",
            ),
            CEMemberContract(
                role=INTERVIEW_CURRENT,
                path="intrvw24/fmli251.csv",
                family="interview_family",
                rows=1,
                columns=9,
                quarter="251",
            ),
            CEMemberContract(
                role=INTERVIEW_CURRENT,
                path="intrvw24/mtbi251.csv",
                family="interview_expenditures",
                rows=1,
                columns=4,
                quarter="251",
            ),
            CEMemberContract(
                role=DIARY_CURRENT,
                path="diary24/fmld241.csv",
                family="diary_family",
                rows=1,
                columns=2,
                quarter="241",
            ),
            CEMemberContract(
                role=DIARY_CURRENT,
                path="diary24/expd241.csv",
                family="diary_expenditures",
                rows=1,
                columns=3,
                quarter="241",
            ),
        ],
        weight_contract=CEWeightContract(
            full_sample="FINLWT21",
            replicate_prefix="WTREP",
            replicate_count=2,
            variance_factor=0.5,
        ),
        benchmark=CEBenchmark(
            ucc="1",
            label="Test tuition",
            source_code="I",
            reference_year=2024,
            expected_mean=40.0,
            expected_standard_error=0.0,
            expected_population_thousands=0.02,
            expected_positive_reporters=2,
            rounding_tolerance=0.0001,
            population_rounding_tolerance=0.0001,
            workbook_sheet="Table 2500",
            workbook_label="Test tuition [I]",
        ),
        dictionary=CEDictionaryContract(
            sheet="Variables",
            required_variables=[
                "NEWID",
                "FINLWT21",
                "WTREP01",
                "WTREP02",
                "QINTRVMO",
                "QINTRVYR",
                "UCC",
                "COST",
                "REF_YR",
            ],
        ),
    )


def _page() -> str:
    return "".join(
        [
            '<a href="/cex/pumd/data/csv/intrvw23.zip">prior</a>',
            '<a href="/cex/pumd/data/csv/intrvw24.zip">current</a>',
            '<a href="/cex/pumd/data/csv/diary24.zip">diary</a>',
            '<a href="/cex/pumd/ce-pumd-interview-diary-dictionary.xlsx">dictionary</a>',
            '<a href="/cex/pumd/errata/errata-2023.htm">errata</a>',
        ]
    )


def _stored(
    role: str,
    relative_path: str,
    *,
    extracted_files: list[str] | None = None,
    documentation: bool = False,
) -> StoredArtifact:
    return StoredArtifact(
        role=role,
        source_url=_url_adapter.validate_python(f"https://www.bls.gov/fixture/{role}"),
        filename=Path(relative_path).name,
        relative_path=relative_path,
        sha256="c" * 64,
        bytes=1,
        extracted_files=extracted_files or [],
        documentation=documentation,
    )


def _write_workbooks(root: Path) -> None:
    benchmark = Workbook()
    sheet = benchmark.active
    assert sheet is not None
    sheet.title = "Table 2500"
    sheet.append(["Test tuition [I]", None])
    sheet.append(["Mean", 40.0])
    sheet.append(["Share", 1.0])
    sheet.append(["SE", 0.0])
    benchmark.save(root / "artifacts/benchmark_workbook/benchmark.xlsx")

    integration = Workbook()
    sheet = integration.active
    assert sheet is not None
    sheet.title = "AllYears_IntStub"
    sheet.append(["title", None, None, None])
    sheet.append(["note", None, None, None])
    sheet.append(["Description", "UCC", "y24", "Level"])
    sheet.append(["Test tuition", "1", "I", 5])
    integration.save(root / "artifacts/integration_workbook/integration.xlsx")

    dictionary = Workbook()
    sheet = dictionary.active
    assert sheet is not None
    sheet.title = "Variables"
    sheet.append(["Survey", "File", "Variable Name"])
    for variable in _definition().dictionary.required_variables:
        sheet.append(["TEST", "TEST", variable])
    dictionary.save(root / "artifacts/dictionary/dictionary.xlsx")


def _fixture(root: Path, *, changed_replicate: bool = False) -> list[StoredArtifact]:
    paths = {
        INTERVIEW_PRIOR: [
            "extracted/interview_prior/intrvw23/fmli241.csv",
            "extracted/interview_prior/intrvw23/mtbi241.csv",
        ],
        INTERVIEW_CURRENT: [
            "extracted/interview_current/intrvw24/fmli251.csv",
            "extracted/interview_current/intrvw24/mtbi251.csv",
        ],
        DIARY_CURRENT: [
            "extracted/diary_current/diary24/fmld241.csv",
            "extracted/diary_current/diary24/expd241.csv",
        ],
    }
    for values in paths.values():
        for relative in values:
            (root / relative).parent.mkdir(parents=True, exist_ok=True)
    replicate_two = 40.0
    second_cost = 20.0 if changed_replicate else 10.0
    pd.DataFrame(
        {
            "NEWID": [1],
            "FINLWT21": [40.0],
            "QINTRVMO": [12],
            "QINTRVYR": [2024],
            "WTREP01": [40.0],
            "WTREP02": [None],
            "AGE_REF": [50],
            "AGE_REF_": ["A"],
            "SOURCE": ["prior"],
        }
    ).to_csv(root / paths[INTERVIEW_PRIOR][0], index=False)
    pd.DataFrame({"NEWID": [1], "COST": [10.0], "UCC": [1], "REF_YR": [2024]}).to_csv(
        root / paths[INTERVIEW_PRIOR][1], index=False
    )
    pd.DataFrame(
        {
            "NEWID": [2],
            "FINLWT21": [40.0],
            "QINTRVMO": [1],
            "QINTRVYR": [2025],
            "WTREP01": [None],
            "WTREP02": [replicate_two],
            "AGE_REF": [40],
            "AGE_REF_": ["B"],
            "SOURCE": ["current"],
        }
    ).to_csv(root / paths[INTERVIEW_CURRENT][0], index=False)
    pd.DataFrame({"NEWID": [2], "COST": [second_cost], "UCC": [1], "REF_YR": [2024]}).to_csv(
        root / paths[INTERVIEW_CURRENT][1], index=False
    )
    pd.DataFrame({"NEWID": [3], "FINLWT21": [20.0]}).to_csv(
        root / paths[DIARY_CURRENT][0], index=False
    )
    pd.DataFrame({"NEWID": [3], "UCC": [2], "COST": [5.0]}).to_csv(
        root / paths[DIARY_CURRENT][1], index=False
    )

    for directory in [
        "artifacts/benchmark_workbook",
        "artifacts/integration_workbook",
        "artifacts/dictionary",
        "artifacts/sample_r",
        "artifacts/getting_started",
        "artifacts/prior_errata",
    ]:
        (root / directory).mkdir(parents=True, exist_ok=True)
    _write_workbooks(root)
    sample = root / "extracted/sample_r/calendar_year_estimate_ucc.R"
    sample.parent.mkdir(parents=True, exist_ok=True)
    sample.write_text("FINLWT21 QINTRVMO QINTRVYR popwt")
    (root / "artifacts/sample_r/sample.zip").write_bytes(b"fixture")
    (root / "artifacts/getting_started/guide.htm").write_text("<main>guide</main>")
    (root / "artifacts/prior_errata/errata.htm").write_text("<main>errata</main>")

    artifacts = [
        _stored(
            role,
            f"artifacts/{role}/{role}.zip",
            extracted_files=extracted,
        )
        for role, extracted in paths.items()
    ]
    artifacts.extend(
        [
            _stored(DICTIONARY, "artifacts/dictionary/dictionary.xlsx"),
            _stored(GETTING_STARTED, "artifacts/getting_started/guide.htm", documentation=True),
            _stored(
                INTEGRATION_WORKBOOK,
                "artifacts/integration_workbook/integration.xlsx",
            ),
            _stored(BENCHMARK_WORKBOOK, "artifacts/benchmark_workbook/benchmark.xlsx"),
            _stored(
                SAMPLE_R,
                "artifacts/sample_r/sample.zip",
                extracted_files=["extracted/sample_r/calendar_year_estimate_ucc.R"],
            ),
            _stored(PRIOR_ERRATA, "artifacts/prior_errata/errata.htm", documentation=True),
        ]
    )
    return artifacts


def _adapter() -> CEAdapter:
    client = httpx.Client(
        transport=httpx.MockTransport(lambda _: httpx.Response(200, text=_page()))
    )
    return CEAdapter(definition=_definition(), client=client)


def test_ce_discovers_exact_csv_and_document_roles() -> None:
    adapter = _adapter()
    release = adapter.discover(2024)
    roles = {str(getattr(item.role, "value", item.role)) for item in release.artifacts}
    assert roles == PAGE_ROLES | STATIC_ROLES
    assert release.source_metadata["adjacent_archive_required"] is True
    adapter.close()


def test_ce_validates_brr_benchmark_and_normalizes(tmp_path: Path) -> None:
    adapter = _adapter()
    release = adapter.discover(2024)
    artifacts = _fixture(tmp_path)
    result = adapter.validate_release(tmp_path, release, artifacts)
    assert result.passed, {key: value for key, value in result.checks.items() if not value}
    outputs = adapter.normalize_release(tmp_path, release, artifacts)
    assert {path.name for path in outputs} == {
        "interview_family.parquet",
        "interview_expenditures.parquet",
        "diary_family.parquet",
        "diary_expenditures.parquet",
    }
    assert pq.ParquetFile(tmp_path / "normalized/interview_family.parquet").metadata.num_rows == 2
    adapter.close()


def test_ce_rejects_replicate_benchmark_mismatch(tmp_path: Path) -> None:
    adapter = _adapter()
    release = adapter.discover(2024)
    artifacts = _fixture(tmp_path, changed_replicate=True)
    result = adapter.validate_release(tmp_path, release, artifacts)
    assert not result.passed
    assert not result.checks["ce_benchmark_standard_error"]
    adapter.close()
